# -*- coding:utf-8 -*-

from subprocess import PIPE
from selenium import webdriver
from time import sleep
from openpyxl import workbook , load_workbook
import os ,ytFuntion ,re ,subprocess ,time

#1.1修正為生產下載位置

def downloadAPK():
   countTime = 0
   if os.path.isfile(appName + ".apk"):    #先確認檔案是否存在
      os.remove(appName +".apk")
   elif os.path.isfile(appName + ".apk.crdownload"):
      os.remove(appName + ".apk.crdownload")
   control_Web.webDriver.get(re.sub("/dev" ,"" ,url)) #生產下載位置
   while True:
      if os.path.isfile(appName +".apk"):
         return " "
         break
      elif countTime == 10:
         return appName
         break        
      else:
         sleep(1)
         countTime = countTime + 1
         

def appspilt():
   if 1 == 1:
      if os.path.isfile(appName + ".apk"):
         output = subprocess.Popen("aapt dump badging " + os.getcwd() + "\\"  +appName +".apk" ,stdout=PIPE ,stderr=PIPE ,stdin=PIPE ,shell = True).stdout.read().decode('utf-8')
      elif os.path.isfile(appName + ".apk.crdownload"):
         output = subprocess.Popen("aapt dump badging " + os.getcwd() + "\\"  +appName +".apk.crdownload" ,stdout=PIPE ,stderr=PIPE ,stdin=PIPE ,shell = True).stdout.read().decode('utf-8')
      j = len(sheetApp["B"])
      try:
         packagename = re.search("package: name='(\S+)'" ,output).group(1)
         versionCode = re.search("versionCode='(\S+)'" ,output).group(1)
         versionName = re.search("versionName='(\S+)'" ,output).group(1)
         launchable_activity_name = re.search("launchable-activity: name='(\S+)'" ,output).group(1)
         sheetApp["B" + str(j+1)].value = url_Number #不存在則更新
         sheetApp["C" + str(j+1)].value = url_name.strip()
         sheetApp["D" + str(j+1)].value = packagename
         sheetApp["E" + str(j+1)].value = versionCode
         sheetApp["F" + str(j+1)].value = versionName
         sheetApp["G" + str(j+1)].value = launchable_activity_name
         wb.save("生產.xlsx") 
         return " "
      except:
         print(appName + "APP分析失敗。")

print("APP下載。")

wb = load_workbook("設定.xlsx")

sheet = wb["帳號"] # 獲取一張表
for i in range(1 ,len(sheet["B"])+1):
   if str(sheet["B" + str(i)].value).strip() == "google":
      google_Account_sitting = str(sheet["C" + str(i)].value).strip() #帳號相關
      google_Password_sitting = str(sheet["D" + str(i)].value).strip()
   if str(sheet["B" + str(i)].value).strip() == "總控":
      control_Account = str(sheet["C" + str(i)].value).strip()
      control_Password = str(sheet["D" + str(i)].value).strip()
      
sheet = wb["url"] # 獲取一張表
for i in range(1 ,len(sheet["B"])+1):
   if str(sheet["B" + str(i)].value).strip() == "總控":
      control_Url = str(sheet["D" + str(i)].value).strip() #URL相關
   if str(sheet["B" + str(i)].value).strip() == "總控站點":
      control_Sitting_url = str(sheet["D" + str(i)].value).strip()
   if str(sheet["B" + str(i)].value).strip() == "APKIOS":
      google_Driver_url = str(sheet["D" + str(i)].value).strip()

testdayFile = time.strftime("%y_%m_%d") + "_生產"
   
if not os.path.exists(testdayFile):    #先確認資料夾是否存在
    os.makedirs(testdayFile)

ogCwd = os.getcwd()
os.chdir(testdayFile)

options = webdriver.ChromeOptions()
prefs = {"profile.default_content_settings.popups": 0, "download.default_directory": os.getcwd()}
options.add_experimental_option("prefs", prefs)

os.chdir(ogCwd)

control_Web = ytFuntion.test_web(webdriver.Chrome(executable_path='chromedriver.exe', chrome_options=options))
control_Web.webDriver.get(google_Driver_url) #目標網址

os.chdir(testdayFile)

if os.path.isfile("DEV-ipa%2Fapk 下載位置.xlsx"):    #先確認檔案是否存在
   os.remove("DEV-ipa%2Fapk 下載位置.xlsx")

control_Web.elementSendKeys("identifierId" ,1 ,text = google_Account_sitting)#帳號
control_Web.elementClick("div[id='identifierNext'] content[class='CwaK9'] span[class='RveJvd snByac']" ,6)

control_Web.elementSendKeys("input[type=password]" ,6 ,text = google_Password_sitting)#密碼
control_Web.elementClick("div[id='passwordNext'] content[class='CwaK9'] span[class='RveJvd snByac']" ,6)

sleep(5)
if not os.path.isfile("DEV-ipa%2Fapk 下載位置.xlsx"):
   input("確認google有成功登入後,請按ENTER繼續")

wbDownload = load_workbook("DEV-ipa%2Fapk 下載位置.xlsx",data_only=True) # 打開一個活頁薄
wbDownload.save(r"DEV-ipa%2Fapk 下載位置.xlsx") #Excel公式處理
wbDownload = load_workbook("DEV-ipa%2Fapk 下載位置.xlsx")
sheetDownload = wbDownload["下載地點"] # 獲取一張表

downloadAPKFail = []

del wb["APP"]
wb.create_sheet(index=3, title="APP")

sheetApp = wb["APP"]
sheetApp["B2"].value = "編號"
sheetApp["C2"].value = "名稱"
sheetApp["D2"].value = "packagename"
sheetApp["E2"].value = "versionCode"
sheetApp["F2"].value = "versionName"
sheetApp["G2"].value = "launchable_activity_name"

for i in range(1 ,len(sheetDownload["A"])+1):
   try:
      if sheetDownload["A" + str(i)].value.strip() == "A001":
         startNumber = i
   except:
      pass

for i in range(int(startNumber) ,len(sheetDownload["A"])+1):
   if sheetDownload["A" + str(i)].value == None:
      break
   url_Number = sheetDownload["A" + str(i)].value.strip()
   url = sheetDownload["E" + str(i)].value.strip()
   url_name = sheetDownload["B" + str(i)].value
   appName = sheetDownload["C" + str(i)].value
   print(url.strip())
   print(url_name.strip())

   downloadAPKFail.append(downloadAPK())
   appspilt()
   
os.chdir(ogCwd)
wb.save("生產.xlsx")
#control_Web.webDriver.quit()
print(downloadAPKFail)
